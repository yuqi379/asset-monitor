#!/usr/bin/env python3
"""
关键词扩展器
- 读取Excel关键词列表
- 自动生成Hunter组合搜索语法
- 支持字段：title, web.body, cert, domain, url
"""

import os
import base64
from typing import List, Dict, Any, Optional

def expand_keyword(keyword: str) -> str:
    """
    将关键词扩展为Hunter查询语法
    只使用 title 和 body 两个字段，用 || 连接
    示例: "登录" -> 'title="登录" || body="登录"'
    """
    return f'title="{keyword}" || body="{keyword}"'


def build_unit_query(unit_name: str) -> str:
    """
    构建项目单位查询语法
    示例: "北京奇安信" -> 'icp.name="北京奇安信"'
    """
    return f'icp.name="{unit_name}"'


def build_domain_query(domain: str) -> str:
    """
    构建域名查询语法
    示例: "qianxin.com" -> 'domain="qianxin.com"'
    """
    return f'domain="{domain}"'


def encode_query(query: str) -> str:
    """将查询语法编码为base64url（Hunter API格式）"""
    return base64.urlsafe_b64encode(query.encode('utf-8')).decode('utf-8').rstrip('=')


def parse_excel_keywords(file_path: str) -> List[Dict[str, Any]]:
    """
    读取Excel文件，解析关键词列表
    Excel格式：
    - 第一列：关键词（选填）
    - 第二列：项目单位（选填）
    - 第三列：域名（选填）
    - 第四列：任务名称（选填，默认用关键词作为名称）
    - 第五列：启用状态（选填，0=禁用，1=启用，默认1）

    每行根据非空字段拆成1~3条独立查询，所有查询共享同一个task_name。

    返回: [{'keyword': '...', 'task_name': '...', 'enabled': True, 'search_query': '...', 'encoded_query': '...'}, ...]
    """
    import openpyxl

    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel文件不存在: {file_path}")

    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    tasks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        keyword = row[0]
        unit_name = row[1] if len(row) > 1 else None
        domain = row[2] if len(row) > 2 else None
        task_name = row[3] if len(row) > 3 else (keyword if keyword else (unit_name if unit_name else domain))
        enabled = True if len(row) <= 4 or row[4] is None else (int(row[4]) == 1)

        # 根据非空字段构建查询
        search_queries = []
        if keyword and str(keyword).strip():
            search_queries.append({
                'query_type': 'keyword',
                'query': expand_keyword(str(keyword).strip()),
            })
        if unit_name and str(unit_name).strip():
            search_queries.append({
                'query_type': 'unit',
                'query': build_unit_query(str(unit_name).strip()),
            })
        if domain and str(domain).strip():
            search_queries.append({
                'query_type': 'domain',
                'query': build_domain_query(str(domain).strip()),
            })

        if not search_queries:
            continue

        task_name_str = str(task_name).strip() if task_name else f"task_{row_idx}"

        for sq in search_queries:
            tasks.append({
                'keyword': str(keyword).strip() if keyword else '',
                'unit_name': str(unit_name).strip() if unit_name else '',
                'domain': str(domain).strip() if domain else '',
                'task_name': task_name_str,
                'query_type': sq['query_type'],
                'enabled': enabled,
                'search_query': sq['query'],
                'encoded_query': encode_query(sq['query']),
            })

    return tasks


def export_keywords_to_excel(tasks: List[Dict[str, Any]], output_path: str):
    """
    将关键词列表导出为Excel（用于模板）
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "关键词列表"

    # 表头
    headers = ['关键词', '项目单位', '域名', '任务名称', '启用状态', '搜索语法（自动生成）']
    ws.append(headers)

    # 按task_name去重，合并同一任务的多个查询语法
    task_map = {}
    for task in tasks:
        tn = task['task_name']
        if tn not in task_map:
            task_map[tn] = {
                'keyword': task['keyword'],
                'unit_name': task['unit_name'],
                'domain': task['domain'],
                'task_name': task['task_name'],
                'enabled': task['enabled'],
                'queries': []
            }
        task_map[tn]['queries'].append(task['search_query'])

    for tn, data in task_map.items():
        ws.append([
            data['keyword'],
            data['unit_name'],
            data['domain'],
            data['task_name'],
            1 if data['enabled'] else 0,
            ' | '.join(data['queries']),
        ])

    # 设置列宽
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 80

    wb.save(output_path)
    print(f"关键词模板已导出: {output_path}")


if __name__ == "__main__":
    # Demo 测试
    print("关键词扩展器 Demo")
    print("=" * 60)

    # 测试单个关键词扩展
    keyword = "电力交易系统"
    query = expand_keyword(keyword)
    print(f"\n关键词: {keyword}")
    print(f"扩展后: {query}")
    print(f"编码后: {encode_query(query)}")

    # 测试Excel导出模板
    demo_tasks = [
        {'keyword': '电力交易系统', 'task_name': '电力交易系统查询', 'enabled': True, 'search_query': ''},
        {'keyword': '医院', 'task_name': '医院系统查询', 'enabled': True, 'search_query': ''},
        {'keyword': '学校', 'task_name': '学校系统查询', 'enabled': False, 'search_query': ''},
    ]
    for t in demo_tasks:
        t['search_query'] = expand_keyword(t['keyword'])
    output = os.path.join(os.path.dirname(__file__), '..', 'keywords_template.xlsx')
    export_keywords_to_excel(demo_tasks, output)
